from flask import Response
import csv
import io
from datetime import datetime

def parse_csv_file(file):
    """Parse un fichier CSV et retourne les donnees"""
    stream = io.StringIO(file.stream.read().decode("UTF8"), newline=None)
    csv_data = csv.DictReader(stream)
    data = []
    for row in csv_data:
        data.append(row)
    return data

def validate_import_data(data, required_fields):
    """Valide les donnees importees"""
    errors = []
    
    if not data:
        return False, ["Le fichier est vide"]
    
    first_row = data[0]
    missing_fields = [field for field in required_fields if field not in first_row.keys()]
    
    if missing_fields:
        errors.append(f"Champs manquants: {', '.join(missing_fields)}")
        return False, errors
    
    return True, []

def export_to_csv(data, filename, headers=None):
    """Exporter des données vers CSV"""
    si = io.StringIO()
    
    if not data:
        return Response('', mimetype='text/csv')
    
    if headers is None:
        headers = data[0].keys() if data else []
    
    writer = csv.DictWriter(si, fieldnames=headers)
    writer.writeheader()
    
    for row in data:
        writer.writerow(row)
    
    output = si.getvalue()
    si.close()
    
    response = Response(output, mimetype='text/csv')
    response.headers['Content-Disposition'] = f'attachment; filename={filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    
    return response

def export_to_excel(data, filename, headers=None, sheet_name='Sheet1'):
    """Exporter des données vers Excel"""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        
        if not data:
            return Response('', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        if headers is None:
            headers = list(data[0].keys()) if data else []
        
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        
        for row_num, row_data in enumerate(data, 2):
            for col_num, header in enumerate(headers, 1):
                ws.cell(row=row_num, column=col_num, value=row_data.get(header, ''))
        
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 15
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = Response(
            output.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response.headers['Content-Disposition'] = f'attachment; filename={filename}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except ImportError:
        return export_to_csv(data, filename, headers)
