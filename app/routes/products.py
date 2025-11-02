from flask import Blueprint, render_template, request, jsonify, flash, redirect, url_for
from app.decorators import require_permission
from flask_login import login_required, current_user
from app.models import db, Product, ProductBatch, Audit, Pharmacy
from app.pharmacy_utils import filter_by_pharmacy, get_accessible_pharmacies, is_admin
from app.export_utils import export_to_csv, export_to_excel, parse_csv_file, validate_import_data
from datetime import datetime

products_bp = Blueprint('products', __name__, url_prefix='/products')

@products_bp.route('/quick-view/<int:id>')
@login_required
def quick_view(id):
    """Vue rapide d'un produit en modal"""
    product = Product.query.get_or_404(id)
    return jsonify({
        'id': product.id,
        'name': product.name,
        'barcode': product.barcode,
        'category': product.category,
        'description': product.description,
        'purchase_price': product.purchase_price,
        'selling_price': product.selling_price,
        'wholesale_price': product.wholesale_price,
        'stock_quantity': product.stock_quantity,
        'min_stock_level': product.min_stock_level,
        'is_low_stock': product.is_low_stock,
        'expiry_date': product.expiry_date.strftime('%d/%m/%Y') if product.expiry_date else None,
        'is_expired': product.is_expired,
        'manufacturer': product.manufacturer,
        'supplier': product.supplier,
        'unit': product.unit,
        'pharmacy_name': product.pharmacy.name if product.pharmacy else 'N/A'
    })

@products_bp.route('/')
@require_permission('manage_products')
def index():
    page = request.args.get('page', 1, type=int)
    search = request.args.get('search', '')
    pharmacy_filter = request.args.get('pharmacy_id', 'all')
    
    query = Product.query.filter_by(is_active=True)
    
    # Filtrage par pharmacie selon les permissions
    query = filter_by_pharmacy(query, Product, pharmacy_filter)
    
    if search:
        query = query.filter(
            db.or_(
                Product.name.ilike(f'%{search}%'),
                Product.barcode.ilike(f'%{search}%'),
                Product.category.ilike(f'%{search}%')
            )
        )
    
    products = query.order_by(Product.created_at.desc()).paginate(
        page=page, per_page=6, error_out=False
    )
    
    # Récupérer les pharmacies accessibles pour le filtre
    pharmacies = get_accessible_pharmacies()
    
    return render_template('products/index.html', 
                         products=products, 
                         search=search,
                         pharmacies=pharmacies,
                         pharmacy_filter=pharmacy_filter)

@products_bp.route('/add', methods=['GET', 'POST'])
@require_permission('manage_products')
def add():
    if request.method == 'POST':
        try:
            primary_pharmacy = current_user.get_primary_pharmacy()
            
            product = Product(
                name=request.form.get('name'),
                description=request.form.get('description'),
                barcode=request.form.get('barcode'),
                category=request.form.get('category'),
                unit=request.form.get('unit', 'piece'),
                purchase_price=float(request.form.get('purchase_price', 0)),
                selling_price=float(request.form.get('selling_price', 0)),
                wholesale_price=float(request.form.get('wholesale_price', 0)),
                stock_quantity=int(request.form.get('stock_quantity', 0)),
                min_stock_level=int(request.form.get('min_stock_level', 10)),
                manufacturer=request.form.get('manufacturer'),
                supplier=request.form.get('supplier'),
                pharmacy_id=primary_pharmacy.id if primary_pharmacy else None
            )
            
            expiry_date_str = request.form.get('expiry_date')
            if expiry_date_str:
                product.expiry_date = datetime.strptime(expiry_date_str, '%Y-%m-%d').date()
            
            db.session.add(product)
            db.session.flush()
            
            audit = Audit(
                user_id=current_user.id,
                action='create_product',
                entity_type='product',
                entity_id=product.id,
                details=f'Produit créé: {product.name}',
                ip_address=request.remote_addr
            )
            db.session.add(audit)
            
            db.session.commit()
            
            flash('Produit ajouté avec succès!', 'success')
            return redirect(url_for('products.index'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')
    
    return render_template('products/add.html')

@products_bp.route('/add-multiple', methods=['GET', 'POST'])
@require_permission('manage_products')
def add_multiple():
    """Ajouter plusieurs produits à la fois"""
    if request.method == 'POST':
        try:
            primary_pharmacy = current_user.get_primary_pharmacy()
            
            # Récupérer les produits depuis le formulaire
            products_data = request.form.getlist('products[]')
            added_count = 0
            errors = []
            
            # Parser les données JSON de chaque produit
            import json
            for product_json in products_data:
                try:
                    product_data = json.loads(product_json)
                    
                    # Validation des champs requis
                    if not product_data.get('name'):
                        errors.append(f"Produit sans nom ignoré")
                        continue
                    
                    # Créer le produit
                    product = Product(
                        name=product_data.get('name', '').strip(),
                        description=product_data.get('description', '').strip(),
                        barcode=product_data.get('barcode', '').strip() or None,
                        category=product_data.get('category', '').strip() or None,
                        unit=product_data.get('unit', 'piece'),
                        purchase_price=float(product_data.get('purchase_price', 0) or 0),
                        selling_price=float(product_data.get('selling_price', 0) or 0),
                        wholesale_price=float(product_data.get('wholesale_price', 0) or 0),
                        stock_quantity=int(product_data.get('stock_quantity', 0) or 0),
                        min_stock_level=int(product_data.get('min_stock_level', 10) or 10),
                        manufacturer=product_data.get('manufacturer', '').strip() or None,
                        supplier=product_data.get('supplier', '').strip() or None,
                        pharmacy_id=primary_pharmacy.id if primary_pharmacy else None
                    )
                    
                    # Gérer la date d'expiration
                    expiry_date_str = product_data.get('expiry_date')
                    if expiry_date_str:
                        try:
                            product.expiry_date = datetime.strptime(expiry_date_str, '%Y-%m-%d').date()
                        except ValueError:
                            pass
                    
                    db.session.add(product)
                    db.session.flush()
                    
                    # Audit
                    audit = Audit(
                        user_id=current_user.id,
                        action='create_product',
                        entity_type='product',
                        entity_id=product.id,
                        details=f'Produit créé (ajout multiple): {product.name}',
                        ip_address=request.remote_addr
                    )
                    db.session.add(audit)
                    
                    added_count += 1
                    
                except json.JSONDecodeError:
                    errors.append("Format de données invalide pour un produit")
                except Exception as e:
                    errors.append(f"Erreur lors de l'ajout d'un produit: {str(e)}")
            
            db.session.commit()
            
            if added_count > 0:
                flash(f'{added_count} produit(s) ajouté(s) avec succès!', 'success')
            if errors:
                for error in errors:
                    flash(error, 'warning')
            
            return redirect(url_for('products.index'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')
    
    return render_template('products/add_multiple.html')

@products_bp.route('/edit/<int:id>', methods=['GET', 'POST'])
@require_permission('manage_products')
def edit(id):
    product = Product.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            product.name = request.form.get('name')
            product.description = request.form.get('description')
            product.barcode = request.form.get('barcode')
            product.category = request.form.get('category')
            product.unit = request.form.get('unit', 'piece')
            product.purchase_price = float(request.form.get('purchase_price', 0))
            product.selling_price = float(request.form.get('selling_price', 0))
            product.wholesale_price = float(request.form.get('wholesale_price', 0))
            product.min_stock_level = int(request.form.get('min_stock_level', 10))
            product.manufacturer = request.form.get('manufacturer')
            product.supplier = request.form.get('supplier')
            
            expiry_date_str = request.form.get('expiry_date')
            if expiry_date_str:
                product.expiry_date = datetime.strptime(expiry_date_str, '%Y-%m-%d').date()
            
            audit = Audit(
                user_id=current_user.id,
                action='update_product',
                entity_type='product',
                entity_id=product.id,
                details=f'Produit modifié: {product.name}',
                ip_address=request.remote_addr
            )
            db.session.add(audit)
            
            db.session.commit()
            
            flash('Produit modifié avec succès!', 'success')
            return redirect(url_for('products.index'))
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur: {str(e)}', 'danger')
    
    return render_template('products/edit.html', product=product)

@products_bp.route('/delete/<int:id>', methods=['POST'])
@require_permission('manage_products')
def delete(id):
    product = Product.query.get_or_404(id)
    
    try:
        product.is_active = False
        
        audit = Audit(
            user_id=current_user.id,
            action='delete_product',
            entity_type='product',
            entity_id=product.id,
            details=f'Produit supprimé: {product.name}',
            ip_address=request.remote_addr
        )
        db.session.add(audit)
        
        db.session.commit()
        flash('Produit supprimé avec succès!', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Erreur: {str(e)}', 'danger')
    
    return redirect(url_for('products.index'))

@products_bp.route('/alerts')
@require_permission('manage_products')
def alerts():
    today = datetime.now().date()
    pharmacy_filter = request.args.get('pharmacy_id', 'all')
    
    # Requête pour le stock faible
    low_stock_query = Product.query.filter(
        Product.is_active == True,
        Product.stock_quantity <= Product.min_stock_level
    )
    low_stock_query = filter_by_pharmacy(low_stock_query, Product, pharmacy_filter)
    low_stock = low_stock_query.all()
    
    # Requête pour les produits expirés
    expired_query = Product.query.filter(
        Product.is_active == True,
        Product.expiry_date < today
    )
    expired_query = filter_by_pharmacy(expired_query, Product, pharmacy_filter)
    expired = expired_query.all()
    
    pharmacies = get_accessible_pharmacies() if is_admin() else []
    
    return render_template('products/alerts.html', 
                         low_stock=low_stock, 
                         expired=expired,
                         pharmacies=pharmacies,
                         pharmacy_filter=pharmacy_filter)

@products_bp.route('/export/<format>')
@require_permission('manage_products')
def export(format):
    """Exporte la liste des produits en CSV ou Excel"""
    pharmacy_filter = request.args.get('pharmacy_id', 'all')
    search = request.args.get('search', '')
    
    query = Product.query.filter_by(is_active=True)
    query = filter_by_pharmacy(query, Product, pharmacy_filter)
    
    if search:
        query = query.filter(
            db.or_(
                Product.name.ilike(f'%{search}%'),
                Product.barcode.ilike(f'%{search}%'),
                Product.category.ilike(f'%{search}%')
            )
        )
    
    products = query.order_by(Product.name).all()
    
    # Préparer les données
    headers = ['ID', 'Nom', 'Code-barres', 'Forme', 'Prix Achat', 'Prix Vente', 
               'Prix Gros', 'Stock', 'Stock Min', 'Unité', 'Fabricant', 'Pharmacie']
    
    data = []
    for p in products:
        pharmacy_name = p.pharmacy.name if p.pharmacy else 'N/A'
        data.append({
            'ID': p.id,
            'Nom': p.name,
            'Code-barres': p.barcode or '',
            'Forme': p.category or '',
            'Prix Achat': p.purchase_price,
            'Prix Vente': p.selling_price,
            'Prix Gros': p.wholesale_price,
            'Stock': p.stock_quantity,
            'Stock Min': p.min_stock_level,
            'Unité': p.unit,
            'Fabricant': p.manufacturer or '',
            'Pharmacie': pharmacy_name
        })
    
    if format == 'csv':
        return export_to_csv(data, 'produits', headers)
    elif format == 'excel':
        return export_to_excel(data, 'produits', headers, 'Produits')
    else:
        flash('Format non supporté', 'danger')
        return redirect(url_for('products.index'))

@products_bp.route('/import/template')
@require_permission('manage_products')
def download_import_template():
    """Télécharge un template Excel bien organisé pour l'import de produits"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from flask import Response
    import io
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Template Import Produits"
    
    # Styles
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    instruction_fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
    instruction_font = Font(bold=True, size=11)
    example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    border_style = Side(style='thin', color='000000')
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    center_alignment = Alignment(horizontal='center', vertical='center')
    left_alignment = Alignment(horizontal='left', vertical='center')
    
    # Ligne 1: Titre
    ws.merge_cells('A1:L1')
    ws['A1'] = 'TEMPLATE D\'IMPORT DE PRODUITS - MARCO PHARMA'
    ws['A1'].font = Font(bold=True, size=14, color="366092")
    ws['A1'].alignment = center_alignment
    ws.row_dimensions[1].height = 25
    
    # Ligne 2: Instructions
    ws.merge_cells('A2:L2')
    ws['A2'] = 'Instructions: Remplissez les colonnes avec les données de vos produits. Les colonnes en rouge sont obligatoires.'
    ws['A2'].font = instruction_font
    ws['A2'].fill = instruction_fill
    ws['A2'].alignment = left_alignment
    ws.row_dimensions[2].height = 20
    
    # Ligne 3: Vide pour espacement
    ws.row_dimensions[3].height = 10
    
    # Ligne 4: En-têtes des colonnes
    headers = [
        ('Nom', True),  # (nom, obligatoire)
        ('Description', False),
        ('Code-barres', False),
        ('Forme', False),
        ('Unité', False),
        ('Prix Achat (USD)', False),
        ('Prix Vente (USD)', True),  # Obligatoire
        ('Prix Gros (USD)', False),
        ('Stock', False),
        ('Stock Min', False),
        ('Fabricant', False),
        ('Fournisseur', False)
    ]
    
    for col_num, (header, is_required) in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header + (' *' if is_required else ''))
        cell.font = header_font
        cell.fill = header_fill if is_required else PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = center_alignment
        cell.border = border
    
    ws.row_dimensions[4].height = 25
    
    # Lignes d'exemple
    examples = [
        ['Paracétamol 500mg', 'Antalgique et antipyrétique', '1234567890123', 'Médicaments', 'boite', '2.50', '5.00', '4.00', '100', '20', 'PharmaCo', 'SupplierX'],
        ['Amoxicilline 250mg', 'Antibiotique à large spectre', '9876543210987', 'Médicaments', 'boite', '3.00', '6.50', '5.50', '50', '15', 'MediLab', 'SupplierY'],
        ['Gants chirurgicaux', 'Gants stériles latex', '5555555555555', 'Consommables', 'paire', '0.50', '1.20', '1.00', '200', '50', 'MedSupply', 'SupplierZ']
    ]
    
    for row_num, example_row in enumerate(examples, 5):
        for col_num, value in enumerate(example_row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.fill = example_fill
            cell.alignment = left_alignment
            cell.border = border
        
        # Note d'exemple
        if row_num == 5:
            note_cell = ws.cell(row=row_num, column=len(headers) + 1, value='← Exemples de données (vous pouvez les supprimer)')
            note_cell.font = Font(italic=True, color="666666", size=9)
    
    # Largeur des colonnes
    column_widths = {
        'A': 25,  # Nom
        'B': 35,  # Description
        'C': 18,  # Code-barres
        'D': 15,  # Forme
        'E': 12,  # Unité
        'F': 18,  # Prix Achat
        'G': 18,  # Prix Vente
        'H': 18,  # Prix Gros
        'I': 12,  # Stock
        'J': 12,  # Stock Min
        'K': 20,  # Fabricant
        'L': 20   # Fournisseur
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    
    # Légende en bas
    legend_row = len(examples) + 6
    ws.merge_cells(f'A{legend_row}:L{legend_row}')
    ws[f'A{legend_row}'] = 'Légende: * = Champ obligatoire | Unité: boite, piece, kg, L, etc. | Prix en USD | Stock en nombres entiers'
    ws[f'A{legend_row}'].font = Font(italic=True, size=9, color="666666")
    ws[f'A{legend_row}'].alignment = left_alignment
    
    # Sauvegarder dans un buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    response = Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response.headers['Content-Disposition'] = 'attachment; filename=template_import_produits.xlsx'
    
    return response

@products_bp.route('/import', methods=['GET', 'POST'])
@require_permission('manage_products')
def import_products():
    """Importe des produits depuis un fichier CSV ou Excel"""
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('Aucun fichier sélectionné', 'danger')
            return redirect(url_for('products.import_products'))
        
        file = request.files['file']
        if file.filename == '':
            flash('Aucun fichier sélectionné', 'danger')
            return redirect(url_for('products.import_products'))
        
        file_ext = file.filename.rsplit('.', 1)[1].lower() if '.' in file.filename else ''
        
        if file_ext not in ['csv', 'xlsx', 'xls']:
            flash('Seuls les fichiers CSV et Excel sont acceptés', 'danger')
            return redirect(url_for('products.import_products'))
        
        try:
            # Parser le fichier
            if file_ext == 'csv':
                data = parse_csv_file(file)
            else:
                # Parser Excel avec gestion d'erreurs améliorée
                try:
                    from openpyxl import load_workbook
                    from openpyxl.utils.exceptions import InvalidFileException
                except ImportError:
                    flash('Le module openpyxl n\'est pas installé. Veuillez l\'installer pour importer des fichiers Excel.', 'danger')
                    return redirect(url_for('products.import_products'))
                
                try:
                    wb = load_workbook(file, read_only=True, data_only=True)
                    ws = wb.active
                except InvalidFileException:
                    flash('Le fichier Excel est corrompu ou dans un format non supporté.', 'danger')
                    return redirect(url_for('products.import_products'))
                except Exception as e:
                    flash(f'Erreur lors de la lecture du fichier Excel: {str(e)}', 'danger')
                    return redirect(url_for('products.import_products'))
                
                # Vérifier que la feuille n'est pas vide
                if ws.max_row == 0:
                    flash('Le fichier Excel est vide.', 'danger')
                    return redirect(url_for('products.import_products'))
                
                # Lire les en-têtes (chercher dans les premières lignes)
                headers = None
                start_row = 1
                try:
                    for row_idx in range(1, min(10, ws.max_row + 1)):
                        row_values = [cell.value for cell in ws[row_idx]]
                        if any(val and ('Nom' in str(val) or 'nom' in str(val)) for val in row_values):
                            headers = [str(cell.value).strip() for cell in ws[row_idx] if cell.value]
                            start_row = row_idx + 1
                            break
                except Exception as e:
                    flash(f'Erreur lors de la lecture des en-têtes: {str(e)}', 'danger')
                    return redirect(url_for('products.import_products'))
                
                if not headers:
                    flash('Format de fichier Excel invalide. Veuillez utiliser le template fourni.', 'danger')
                    return redirect(url_for('products.import_products'))
                
                # Mapping des noms de colonnes (normalisation)
                column_mapping = {
                    'Nom': 'Nom',
                    'nom': 'Nom',
                    'Description': 'Description',
                    'description': 'Description',
                    'Code-barres': 'Code-barres',
                    'Code barres': 'Code-barres',
                    'Forme': 'Forme',
                    'forme': 'Forme',
                    'Unité': 'Unité',
                    'Unite': 'Unité',
                    'Prix Achat (USD)': 'Prix Achat',
                    'Prix Achat': 'Prix Achat',
                    'Prix Vente (USD)': 'Prix Vente',
                    'Prix Vente': 'Prix Vente',
                    'Prix Gros (USD)': 'Prix Gros',
                    'Prix Gros': 'Prix Gros',
                    'Stock': 'Stock',
                    'Stock Min': 'Stock Min',
                    'Fabricant': 'Fabricant',
                    'Fournisseur': 'Fournisseur'
                }
                
                # Normaliser les en-têtes
                normalized_headers = []
                for header in headers:
                    header_clean = header.replace(' *', '').strip()
                    normalized_headers.append(column_mapping.get(header_clean, header_clean))
                
                # Lire les données avec gestion d'erreurs
                data = []
                try:
                    for row in ws.iter_rows(min_row=start_row, values_only=False):
                        try:
                            if not any(cell.value for cell in row):
                                continue
                            
                            row_dict = {}
                            for col_idx, original_header in enumerate(headers):
                                if col_idx < len(normalized_headers):
                                    normalized_header = normalized_headers[col_idx]
                                    if col_idx < len(row):
                                        cell_value = row[col_idx].value
                                        if cell_value is not None:
                                            # Convertir les nombres en chaînes pour la cohérence
                                            if isinstance(cell_value, (int, float)):
                                                row_dict[normalized_header] = str(cell_value)
                                            else:
                                                row_dict[normalized_header] = str(cell_value).strip()
                                        else:
                                            row_dict[normalized_header] = ''
                            
                            if any(row_dict.values()):
                                data.append(row_dict)
                        except Exception as row_error:
                            # Continuer avec la ligne suivante en cas d'erreur
                            continue
                except Exception as e:
                    flash(f'Erreur lors de la lecture des données Excel: {str(e)}', 'danger')
                    return redirect(url_for('products.import_products'))
                finally:
                    # Fermer le workbook proprement
                    try:
                        wb.close()
                    except:
                        pass
            
            if data is None:
                return redirect(url_for('products.import_products'))
            
            # Valider les données
            required_fields = ['Nom', 'Prix Vente']
            valid, errors = validate_import_data(data, required_fields)
            
            if not valid:
                for error in errors:
                    flash(error, 'danger')
                return redirect(url_for('products.import_products'))
            
            # Importer les produits
            primary_pharmacy = current_user.get_primary_pharmacy()
            imported_count = 0
            
            for row in data:
                try:
                    # Validation des champs numériques avec gestion d'erreurs
                    try:
                        purchase_price = float(row.get('Prix Achat', 0) or 0)
                    except (ValueError, TypeError):
                        purchase_price = 0.0
                    
                    try:
                        selling_price = float(row.get('Prix Vente', 0) or 0)
                        if selling_price <= 0:
                            raise ValueError('Le prix de vente doit être supérieur à 0')
                    except (ValueError, TypeError) as e:
                        flash(f'Erreur prix ligne "{row.get("Nom", "?")}": {str(e)}', 'warning')
                        continue
                    
                    try:
                        wholesale_price = float(row.get('Prix Gros', 0) or 0)
                    except (ValueError, TypeError):
                        wholesale_price = 0.0
                    
                    try:
                        stock_quantity = int(row.get('Stock', 0) or 0)
                    except (ValueError, TypeError):
                        stock_quantity = 0
                    
                    try:
                        min_stock_level = int(row.get('Stock Min', 10) or 10)
                    except (ValueError, TypeError):
                        min_stock_level = 10
                    
                    product = Product(
                        name=row['Nom'],
                        description=row.get('Description', ''),
                        barcode=row.get('Code-barres', ''),
                        category=row.get('Forme', ''),
                        unit=row.get('Unité', 'piece'),
                        purchase_price=purchase_price,
                        selling_price=selling_price,
                        wholesale_price=wholesale_price,
                        stock_quantity=stock_quantity,
                        min_stock_level=min_stock_level,
                        manufacturer=row.get('Fabricant', ''),
                        supplier=row.get('Fournisseur', ''),
                        pharmacy_id=primary_pharmacy.id if primary_pharmacy else None
                    )
                    
                    db.session.add(product)
                    imported_count += 1
                except KeyError as e:
                    flash(f'Champ manquant ligne "{row.get("Nom", "?")}": {str(e)}', 'warning')
                    continue
                except Exception as e:
                    flash(f'Erreur ligne "{row.get("Nom", "?")}": {str(e)}', 'warning')
                    continue
            
            # Audit
            audit = Audit(
                user_id=current_user.id,
                action='import_products',
                entity_type='product',
                details=f'{imported_count} produits importés',
                ip_address=request.remote_addr
            )
            db.session.add(audit)
            
            db.session.commit()
            flash(f'{imported_count} produits importés avec succès!', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur lors de l\'import: {str(e)}', 'danger')
        
        return redirect(url_for('products.index'))
    
    return render_template('products/import.html')

@products_bp.route('/toggle-status/<int:id>', methods=['POST'])
@require_permission('manage_products')
def toggle_status(id):
    product = Product.query.get_or_404(id)
    product.is_active = not product.is_active
    
    audit = Audit(
        user_id=current_user.id,
        action='toggle_product_status',
        entity_type='product',
        entity_id=product.id,
        details=f'Statut produit changé: {product.name} -> {"Actif" if product.is_active else "Inactif"}',
        ip_address=request.remote_addr
    )
    db.session.add(audit)
    
    db.session.commit()
    
    flash(f'Produit {"activé" if product.is_active else "désactivé"} avec succès!', 'success')
    return redirect(url_for('products.index'))

@products_bp.route('/export-single/<int:id>')
@require_permission('manage_products')
def export_single(id):
    product = Product.query.get_or_404(id)
    
    # Données à exporter
    export_data = {
        'product': {
            'name': product.name,
            'barcode': product.barcode,
            'category': product.category,
            'unit': product.unit,
            'purchase_price': product.purchase_price,
            'selling_price': product.selling_price,
            'wholesale_price': product.wholesale_price,
            'stock_quantity': product.stock_quantity,
            'min_stock_level': product.min_stock_level,
            'manufacturer': product.manufacturer,
            'supplier': product.supplier,
            'is_active': product.is_active,
            'created_at': product.created_at.isoformat() if product.created_at else None
        },
        'stock_movements': [],
        'sales_summary': {}
    }
    
    # Mouvements de stock
    movements = StockMovement.query.filter_by(product_id=id).order_by(StockMovement.created_at.desc()).limit(10).all()
    for movement in movements:
        export_data['stock_movements'].append({
            'movement_type': movement.movement_type,
            'quantity': movement.quantity,
            'reference': movement.reference,
            'notes': movement.notes,
            'created_at': movement.created_at.isoformat() if movement.created_at else None
        })
    
    # Résumé des ventes
    sales_items = SaleItem.query.filter_by(product_id=id).all()
    total_sold = sum(item.quantity for item in sales_items)
    total_revenue = sum(item.total for item in sales_items)
    
    export_data['sales_summary'] = {
        'total_sold': total_sold,
        'total_revenue': total_revenue
    }
    
    # Créer un fichier JSON
    import json
    from flask import make_response
    
    response = make_response(json.dumps(export_data, indent=2, ensure_ascii=False))
    response.headers['Content-Type'] = 'application/json'
    response.headers['Content-Disposition'] = f'attachment; filename=product_{product.name}_export.json'
    
    return response

@products_bp.route('/duplicate/<int:id>', methods=['GET', 'POST'])
@require_permission('manage_products')
def duplicate(id):
    """Dupliquer un produit"""
    original_product = Product.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            # Créer une copie du produit
            new_product = Product(
                name=request.form.get('name', f"{original_product.name} (Copie)"),
                barcode=request.form.get('barcode', ''),
                category=original_product.category,
                unit=original_product.unit,
                purchase_price=original_product.purchase_price,
                selling_price=original_product.selling_price,
                wholesale_price=original_product.wholesale_price,
                stock_quantity=0,  # Stock à zéro pour la copie
                min_stock_level=original_product.min_stock_level,
                manufacturer=original_product.manufacturer,
                supplier=original_product.supplier,
                description=original_product.description,
                pharmacy_id=original_product.pharmacy_id,
                is_active=True
            )
            
            db.session.add(new_product)
            
            # Audit
            audit = Audit(
                user_id=current_user.id,
                action='duplicate_product',
                entity_type='product',
                entity_id=new_product.id,
                details=f'Produit dupliqué: {original_product.name} -> {new_product.name}',
                ip_address=request.remote_addr
            )
            db.session.add(audit)
            
            db.session.commit()
            flash('Produit dupliqué avec succès!', 'success')
            return redirect(url_for('products.index'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Erreur lors de la duplication: {str(e)}', 'danger')
    
    return render_template('products/duplicate.html', product=original_product)
